import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# --- CONEXIÓN 
conn = pyodbc.connect(
    "DRIVER={SQL Server};"
    "SERVER=LAPTOP-PCT25RMC\\SQLEXPRESS;"
    "DATABASE=ComercialAndina;"
    "Trusted_Connection=yes;"
)
cursor = conn.cursor()

# --- PARÁMETROS 
PERIODO     = "2026-06"   # ← cambia este valor para otro mes
RUC         = "20123456789"
EMPRESA     = "Comercial Andina S.A.C."

MESES_LABEL = {
    "2026-01": "Enero 2026",   "2026-02": "Febrero 2026",
    "2026-03": "Marzo 2026",   "2026-04": "Abril 2026",
    "2026-05": "Mayo 2026",    "2026-06": "Junio 2026",
}

#  QUERY: saldos por cuenta para el periodo 
def obtener_saldos(periodo):
    cursor.execute("""
        SELECT
            d.codigo_cuenta,
            p.nombre_cuenta,
            p.tipo,
            p.naturaleza,
            p.elemento,
            SUM(d.debe)  AS total_debe,
            SUM(d.haber) AS total_haber
        FROM asientos_detalle d
        JOIN asientos_cabecera c ON d.id_asiento = c.id_asiento
        JOIN plan_cuentas p      ON d.codigo_cuenta = p.codigo_cuenta
        WHERE c.periodo = ?
        GROUP BY d.codigo_cuenta, p.nombre_cuenta, p.tipo, p.naturaleza, p.elemento
        ORDER BY d.codigo_cuenta
    """, periodo)
    rows = cursor.fetchall()
    saldos = {}
    for r in rows:
        debe  = float(r.total_debe  or 0)
        haber = float(r.total_haber or 0)
        saldo = debe - haber if r.naturaleza == 'D' else haber - debe
        saldos[r.codigo_cuenta] = {
            "nombre":    r.nombre_cuenta,
            "tipo":      r.tipo,
            "naturaleza":r.naturaleza,
            "elemento":  r.elemento,
            "debe":      debe,
            "haber":     haber,
            "saldo":     saldo,
        }
    return saldos

def suma_cuentas(saldos, codigos):
    return sum(saldos[c]["saldo"] for c in codigos if c in saldos)

# estilo de excel ---- la suma falla pero se llena manual solo dale espacio al texto dentro de l celda
def estilo_titulo(ws, fila, texto, col_fin=4):
    ws.merge_cells(f"A{fila}:{get_column_letter(col_fin)}{fila}")
    c = ws[f"A{fila}"]
    c.value     = texto
    c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")

def estilo_subtitulo(ws, fila, texto, col_fin=4):
    ws.merge_cells(f"A{fila}:{get_column_letter(col_fin)}{fila}")
    c = ws[f"A{fila}"]
    c.value     = texto
    c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="2E75B6")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def estilo_seccion(ws, fila, texto, col_fin=4):
    ws.merge_cells(f"A{fila}:{get_column_letter(col_fin)}{fila}")
    c = ws[f"A{fila}"]
    c.value     = texto
    c.font      = Font(name="Arial", bold=True, size=9, color="1F4E79")
    c.fill      = PatternFill("solid", start_color="D6E4F0")
    c.alignment = Alignment(horizontal="left", indent=1)

def fila_dato(ws, fila, label, valor, sangria=2, negrita=False, col_valor=2):
    ws.cell(fila, 1).value     = label
    ws.cell(fila, 1).font      = Font(name="Arial", bold=negrita, size=9)
    ws.cell(fila, 1).alignment = Alignment(horizontal="left", indent=sangria)
    ws.cell(fila, col_valor).value       = valor if valor != 0 else "-"
    ws.cell(fila, col_valor).font        = Font(name="Arial", bold=negrita, size=9, color="000000")
    ws.cell(fila, col_valor).number_format = '#,##0.00;(#,##0.00);"-"'
    ws.cell(fila, col_valor).alignment   = Alignment(horizontal="right")

def fila_total(ws, fila, label, formula, col_fin=2):
    ws.cell(fila, 1).value     = label
    ws.cell(fila, 1).font      = Font(name="Arial", bold=True, size=9, color="1F4E79")
    ws.cell(fila, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(fila, 1).fill      = PatternFill("solid", start_color="EBF3FB")
    ws.cell(fila, col_fin).value         = formula
    ws.cell(fila, col_fin).font          = Font(name="Arial", bold=True, size=9, color="1F4E79")
    ws.cell(fila, col_fin).number_format = '#,##0.00;(#,##0.00);"-"'
    ws.cell(fila, col_fin).fill          = PatternFill("solid", start_color="EBF3FB")
    ws.cell(fila, col_fin).alignment     = Alignment(horizontal="right")

def linea_borde(ws, fila, cols=4):
    borde = Border(bottom=Side(style="thin", color="2E75B6"))
    for col in range(1, cols + 1):
        ws.cell(fila, col).border = borde

#  HOJA 1: ESTADO DE RESULTADOS 
def hoja_estado_resultados(wb, saldos, periodo):
    ws = wb.create_sheet("Estado de Resultados")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.row_dimensions[1].height     = 28
    ws.row_dimensions[2].height     = 18
    ws.row_dimensions[3].height     = 18

    label_periodo = MESES_LABEL.get(periodo, periodo)

    estilo_titulo(ws, 1, f"ESTADO DE GANANCIAS Y PÉRDIDAS — {label_periodo.upper()}", 2)
    ws["A2"] = f"RUC: {RUC}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws["A3"] = EMPRESA
    ws["A3"].font = Font(name="Arial", size=9, italic=True)

    #  Ingresos 
    r = 5
    estilo_seccion(ws, r, "INGRESOS OPERACIONALES", 2); r += 1
    ventas       = suma_cuentas(saldos, ["70.121"])
    fila_dato(ws, r, "Ventas Netas (ingresos operacionales)", ventas); r += 1
    r_total_ing  = r
    fila_total(ws, r, "Total de Ingresos Brutos", f"=B{r-1}"); r += 1
    linea_borde(ws, r); r += 1

    #  Costos
    estilo_seccion(ws, r, "COSTO", 2); r += 1
    costo        = suma_cuentas(saldos, ["69.121"])
    r_costo      = r
    fila_dato(ws, r, "Costo de Ventas", costo); r += 1
    r_util_bruta = r
    fila_total(ws, r, "Utilidad Bruta", f"=B{r_total_ing}-B{r_costo}"); r += 1
    linea_borde(ws, r); r += 1

    #  Gastos operacionales 
    estilo_seccion(ws, r, "GASTOS OPERACIONALES", 2); r += 1
    g_admin = suma_cuentas(saldos, [
        "62.11","62.14","62.15","62.71","62.72","62.91","627",
        "63.52","63.61","63.63","63.65"
    ])
    g_venta = suma_cuentas(saldos, ["63.111"])
    r_gadmin = r
    fila_dato(ws, r, "Gastos de Administración", g_admin); r += 1
    r_gventa = r
    fila_dato(ws, r, "Gastos de Venta", g_venta); r += 1
    r_util_op = r
    fila_total(ws, r, "Utilidad Operativa", f"=B{r_util_bruta}-B{r_gadmin}-B{r_gventa}"); r += 1
    linea_borde(ws, r); r += 1

    #  Impuesto 
    estilo_seccion(ws, r, "IMPUESTO A LA RENTA", 2); r += 1
    ir = suma_cuentas(saldos, ["40.171"])
    r_ir = r
    fila_dato(ws, r, "Impuesto a la Renta 3ra Categoría", ir); r += 1
    r_util_neta = r
    fila_total(ws, r, "UTILIDAD (PÉRDIDA) NETA DEL EJERCICIO", f"=B{r_util_op}-B{r_ir}"); r += 1

#  HOJA 2: BALANCE GENERAL 
def hoja_balance_general(wb, saldos, periodo):
    ws = wb.create_sheet("Balance General")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.row_dimensions[1].height     = 28

    label_periodo = MESES_LABEL.get(periodo, periodo)
    estilo_titulo(ws, 1, f"BALANCE GENERAL — {label_periodo.upper()}", 4)
    ws["A2"] = f"RUC: {RUC}   |   {EMPRESA}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws.merge_cells("A2:D2")

    # Saldos
    caja       = suma_cuentas(saldos, ["10.1", "10.41.1", "10.41.2"])
    cxc        = suma_cuentas(saldos, ["12.13"])
    exist      = suma_cuentas(saldos, ["20.111"])
    igv_cred   = suma_cuentas(saldos, ["16.73"])
    cxp        = suma_cuentas(saldos, ["42.12"])
    otras_cxp  = suma_cuentas(saldos, [
        "40.111","40.171","40.173","40.31","40.32",
        "41.11","41.14","41.15","41.51","41.7"
    ])
    capital    = suma_cuentas(saldos, ["50.1"])
    utilidades = suma_cuentas(saldos, ["59.11"])

    # ACTIVO columna A/B
    r = 4
    estilo_subtitulo(ws, r, "ACTIVO", 2); r += 1
    estilo_seccion(ws, r, "ACTIVO CORRIENTE", 2); r += 1
    r_caja   = r; fila_dato(ws, r, "Caja y Bancos",                                                   caja,     col_valor=2); r += 1
    r_vn     = r; fila_dato(ws, r, "Valores Negociables",                                             0,        col_valor=2); r += 1
    r_cxcc   = r; fila_dato(ws, r, "Cuentas por Cobrar Comerciales",                                  cxc,      col_valor=2); r += 1
    r_cxcv   = r; fila_dato(ws, r, "Cuentas por Cobrar a Vinculadas",                                 0,        col_valor=2); r += 1
    r_ocxc   = r; fila_dato(ws, r, "Otras Cuentas por Cobrar",                                        igv_cred, col_valor=2); r += 1
    r_exist  = r; fila_dato(ws, r, "Existencias",                                                     exist,    col_valor=2); r += 1
    r_gpa    = r; fila_dato(ws, r, "Gastos Pagados por Anticipado",                                   0,        col_valor=2); r += 1
    r_tac    = r
    fila_total(ws, r, "TOTAL ACTIVO CORRIENTE", f"=B{r_caja}+B{r_vn}+B{r_cxcc}+B{r_cxcv}+B{r_ocxc}+B{r_exist}+B{r_gpa}", col_fin=2); r += 1
    linea_borde(ws, r, 2); r += 1

    estilo_seccion(ws, r, "ACTIVO NO CORRIENTE", 2); r += 1
    r_cxclp  = r; fila_dato(ws, r, "Cuentas por Cobrar a Largo Plazo",                                0, col_valor=2); r += 1
    r_cxcvlp = r; fila_dato(ws, r, "Cuentas por Cobrar a Vinculadas a Largo Plazo",                   0, col_valor=2); r += 1
    r_ocxclp = r; fila_dato(ws, r, "Otras Cuentas por Cobrar a Largo Plazo",                          0, col_valor=2); r += 1
    r_invp   = r; fila_dato(ws, r, "Inversiones Permanentes",                                         0, col_valor=2); r += 1
    r_ime    = r; fila_dato(ws, r, "Inmuebles, Maquinaria y Equipo (neto de depreciacion acumulada)", 0, col_valor=2); r += 1
    r_int    = r; fila_dato(ws, r, "Activos Intangibles (neto de amortizacion acumulada)",            0, col_valor=2); r += 1
    r_ird    = r; fila_dato(ws, r, "Impuesto a la Renta y Participaciones Diferidos Activo",          0, col_valor=2); r += 1
    r_oa     = r; fila_dato(ws, r, "Otros Activos",                                                   0, col_valor=2); r += 1
    r_tanc   = r
    fila_total(ws, r, "TOTAL ACTIVO NO CORRIENTE", f"=B{r_cxclp}+B{r_cxcvlp}+B{r_ocxclp}+B{r_invp}+B{r_ime}+B{r_int}+B{r_ird}+B{r_oa}", col_fin=2); r += 1
    linea_borde(ws, r, 2); r += 1

    r_tact = r
    fila_total(ws, r, "TOTAL ACTIVO", f"=B{r_tac}+B{r_tanc}", col_fin=2); r += 1

    # PASIVO Y PATRIMONIO columna C/D
    r = 4
    estilo_subtitulo(ws, r, "PASIVO Y PATRIMONIO", 4); r += 1
    estilo_seccion(ws, r, "PASIVO CORRIENTE", 4); r += 1
    r_sob    = r; fila_dato(ws, r, "Sobregiros y Pagares Bancarios",                 0,         col_valor=4); r += 1
    r_cxpc   = r; fila_dato(ws, r, "Cuentas por Pagar Comerciales",                 cxp,       col_valor=4); r += 1
    r_cxpv   = r; fila_dato(ws, r, "Cuentas por Pagar a Vinculadas",                0,         col_valor=4); r += 1
    r_ocxp   = r; fila_dato(ws, r, "Otras Cuentas por Pagar",                       otras_cxp, col_valor=4); r += 1
    r_pcdlp  = r; fila_dato(ws, r, "Parte Corriente de las Deudas a Largo Plazo",   0,         col_valor=4); r += 1
    r_tpc    = r
    fila_total(ws, r, "TOTAL PASIVO CORRIENTE", f"=D{r_sob}+D{r_cxpc}+D{r_cxpv}+D{r_ocxp}+D{r_pcdlp}", col_fin=4); r += 1
    linea_borde(ws, r, 4); r += 1

    estilo_seccion(ws, r, "PASIVO NO CORRIENTE", 4); r += 1
    r_dlp    = r; fila_dato(ws, r, "Deudas a Largo Plazo",                               0, col_valor=4); r += 1
    r_cxpvlp = r; fila_dato(ws, r, "Cuentas por Pagar a Vinculadas LP",                  0, col_valor=4); r += 1
    r_ingd   = r; fila_dato(ws, r, "Ingresos Diferidos",                                 0, col_valor=4); r += 1
    r_irdp   = r; fila_dato(ws, r, "Impuesto a la Renta y Participaciones Diferidos Pasivo", 0, col_valor=4); r += 1
    r_tpnc   = r
    fila_total(ws, r, "TOTAL PASIVO NO CORRIENTE", f"=D{r_dlp}+D{r_cxpvlp}+D{r_ingd}+D{r_irdp}", col_fin=4); r += 1
    linea_borde(ws, r, 4); r += 1

    r_tpas = r
    fila_total(ws, r, "TOTAL PASIVO", f"=D{r_tpc}+D{r_tpnc}", col_fin=4); r += 1
    linea_borde(ws, r, 4); r += 1

    fila_dato(ws, r, "Contingencias",        0, col_valor=4); r += 1
    fila_dato(ws, r, "Interes Minoritario",  0, col_valor=4); r += 1
    linea_borde(ws, r, 4); r += 1

    estilo_seccion(ws, r, "PATRIMONIO NETO", 4); r += 1
    r_cap    = r; fila_dato(ws, r, "Capital",                   capital,    col_valor=4); r += 1
    r_capad  = r; fila_dato(ws, r, "Capital Adicional",         0,          col_valor=4); r += 1
    r_ainv   = r; fila_dato(ws, r, "Acciones de Inversion",     0,          col_valor=4); r += 1
    r_exrev  = r; fila_dato(ws, r, "Excedentes de Revaluacion", 0,          col_valor=4); r += 1
    r_resl   = r; fila_dato(ws, r, "Reservas Legales",          0,          col_valor=4); r += 1
    r_ores   = r; fila_dato(ws, r, "Otras Reservas",            0,          col_valor=4); r += 1
    r_util   = r; fila_dato(ws, r, "Resultados Acumulados",     utilidades, col_valor=4); r += 1
    r_tpat   = r
    fila_total(ws, r, "TOTAL PATRIMONIO NETO", f"=D{r_cap}+D{r_capad}+D{r_ainv}+D{r_exrev}+D{r_resl}+D{r_ores}+D{r_util}", col_fin=4); r += 1
    linea_borde(ws, r, 4); r += 1

    r_tpyp = r
    fila_total(ws, r, "TOTAL PASIVO Y PATRIMONIO NETO", f"=D{r_tpas}+D{r_tpat}", col_fin=4); r += 1

# +++ HOJA 3: BALANCE DE COMPROBACIÓN ++++++++++++++++++++++++++++++++++
def hoja_balance_comprobacion(wb, saldos, periodo):
    ws = wb.create_sheet("Balance de Comprobación")
    label_periodo = MESES_LABEL.get(periodo, periodo)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.row_dimensions[1].height     = 28

    estilo_titulo(ws, 1, f"BALANCE DE COMPROBACIÓN — {label_periodo.upper()}", 6)
    ws["A2"] = f"RUC: {RUC}   |   {EMPRESA}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws.merge_cells("A2:F2")

    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    headers = ["Código", "Denominación", "Debe", "Haber", "Saldo Deudor", "Saldo Acreedor"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i)
        c.value     = h
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center")

    r          = 5
    filas_debe = []
    filas_haber= []

    for codigo, datos in sorted(saldos.items()):
        ws.cell(r, 1).value = codigo
        ws.cell(r, 2).value = datos["nombre"]
        ws.cell(r, 3).value = datos["debe"]
        ws.cell(r, 4).value = datos["haber"]

        saldo_d = datos["saldo"] if datos["naturaleza"] == "D" else 0
        saldo_h = datos["saldo"] if datos["naturaleza"] == "H" else 0
        ws.cell(r, 5).value = saldo_d if saldo_d > 0 else None
        ws.cell(r, 6).value = saldo_h if saldo_h > 0 else None

        for col in range(1, 7):
            c = ws.cell(r, col)
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="right" if col > 2 else "left")
            if col > 2:
                c.number_format = '#,##0.00;(#,##0.00);"-"'
            if r % 2 == 0:
                c.fill = PatternFill("solid", start_color="F2F7FC")

        filas_debe.append(r)
        filas_haber.append(r)
        r += 1

    # Totales
    for col, letra in [(3,"C"),(4,"D"),(5,"E"),(6,"F")]:
        c = ws.cell(r, col)
        c.value         = f"=SUM({letra}5:{letra}{r-1})"
        c.font          = Font(name="Arial", bold=True, size=9, color="1F4E79")
        c.number_format = '#,##0.00;(#,##0.00);"-"'
        c.fill          = PatternFill("solid", start_color="EBF3FB")
        c.alignment     = Alignment(horizontal="right")

    ws.cell(r, 1).value = "TOTALES"
    ws.cell(r, 1).font  = Font(name="Arial", bold=True, size=9, color="1F4E79")
    ws.cell(r, 1).fill  = PatternFill("solid", start_color="EBF3FB")

#  HOJA 4: RATIOS FINANCIEROS +++++++++++++++++++++++++++++++++++++++++
def hoja_ratios(wb, saldos, periodo):
    ws = wb.create_sheet("Ratios Financieros")
    label_periodo = MESES_LABEL.get(periodo, periodo)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.row_dimensions[1].height     = 28

    estilo_titulo(ws, 1, f"RATIOS FINANCIEROS — {label_periodo.upper()}", 3)
    ws["A2"] = f"RUC: {RUC}   |   {EMPRESA}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws.merge_cells("A2:C2")

    headers = ["Ratio", "Valor", "Interpretación"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i)
        c.value     = h
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center")

    # Valores base
    activo_cte  = suma_cuentas(saldos, ["10.1","10.41.1","10.41.2","12.13","16.73","20.111"])
    pasivo_cte  = suma_cuentas(saldos, ["42.12","40.111","40.171","40.173","40.31","40.32","41.11","41.14","41.15","41.51","41.7"])
    exist       = suma_cuentas(saldos, ["20.111"])
    ventas      = suma_cuentas(saldos, ["70.121"])
    costo       = suma_cuentas(saldos, ["69.121"])
    g_admin     = suma_cuentas(saldos, ["62.11","62.14","62.15","62.71","62.72","62.91","627","63.52","63.61","63.63","63.65"])
    g_venta     = suma_cuentas(saldos, ["63.111"])
    ir          = suma_cuentas(saldos, ["40.171"])
    activo_total= activo_cte
    util_bruta  = ventas - costo
    util_op     = util_bruta - g_admin - g_venta
    util_neta   = util_op - ir
    cxc         = suma_cuentas(saldos, ["12.13"])

    liq_corriente = round(activo_cte / pasivo_cte, 2)          if pasivo_cte  else 0
    prueba_acida  = round((activo_cte - exist) / pasivo_cte, 2) if pasivo_cte  else 0
    endeudamiento = round(pasivo_cte / activo_total * 100, 2)   if activo_total else 0
    margen_bruto  = round(util_bruta / ventas * 100, 2)         if ventas      else 0
    margen_neto   = round(util_neta  / ventas * 100, 2)         if ventas      else 0
    roa           = round(util_neta  / activo_total * 100, 2)   if activo_total else 0
    rot_cxc       = round(ventas / cxc, 2)                      if cxc         else 0

    ratios = [
        ("LIQUIDEZ",   None, None),
        ("Liquidez Corriente",    liq_corriente,  "Activo Cte / Pasivo Cte  |  >1.5 saludable"),
        ("Prueba Ácida",          prueba_acida,   "(Activo Cte - Existencias) / Pasivo Cte  |  >1.0 ideal"),
        ("SOLVENCIA",  None, None),
        ("Endeudamiento (%)",     endeudamiento,  "Pasivo Total / Activo Total × 100  |  <60% razonable"),
        ("RENTABILIDAD", None, None),
        ("Margen Bruto (%)",      margen_bruto,   "Utilidad Bruta / Ventas × 100"),
        ("Margen Neto (%)",       margen_neto,    "Utilidad Neta / Ventas × 100"),
        ("ROA (%)",               roa,            "Utilidad Neta / Activo Total × 100"),
        ("ACTIVIDAD",  None, None),
        ("Rotación CxC (veces)",  rot_cxc,        "Ventas / CxC  |  cuántas veces se cobra al mes"),
    ]

    r = 5
    for nombre, valor, interp in ratios:
        if valor is None:
            estilo_seccion(ws, r, nombre, 3)
        else:
            ws.cell(r, 1).value     = nombre
            ws.cell(r, 1).font      = Font(name="Arial", size=9)
            ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=2)
            ws.cell(r, 2).value     = valor
            ws.cell(r, 2).font      = Font(name="Arial", bold=True, size=9, color="1F4E79")
            ws.cell(r, 2).number_format = "0.00"
            ws.cell(r, 2).alignment = Alignment(horizontal="center")
            ws.cell(r, 3).value     = interp
            ws.cell(r, 3).font      = Font(name="Arial", size=8, italic=True, color="595959")
            if r % 2 == 0:
                for col in range(1, 4):
                    ws.cell(r, col).fill = PatternFill("solid", start_color="F2F7FC")
        r += 1

#  EJECUCIÓN 
saldos = obtener_saldos(PERIODO)

wb = openpyxl.Workbook()
wb.remove(wb.active)

hoja_estado_resultados(wb, saldos, PERIODO)
hoja_balance_general(wb, saldos, PERIODO)
hoja_balance_comprobacion(wb, saldos, PERIODO)
hoja_ratios(wb, saldos, PERIODO)

nombre_archivo = f"Estados_Financieros_{PERIODO}.xlsx"
wb.save(nombre_archivo)
conn.close()

print(f"✅ Archivo generado: {nombre_archivo}")